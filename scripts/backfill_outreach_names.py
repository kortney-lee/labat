"""
Backfill company_name and fix first_name on outreach_leads from spreadsheets.
Run with: python scripts/backfill_outreach_names.py
"""
import os, subprocess, json, time
import openpyxl
import urllib.request
import urllib.error

PROJECT = "wihy-ai"
COLL    = "outreach_leads"
BASE    = f"https://firestore.googleapis.com/v1/projects/{PROJECT}/databases/(default)/documents"

def get_token():
    import shutil
    gcloud = shutil.which("gcloud") or r"C:\Program Files (x86)\Google\Cloud SDK\google-cloud-sdk\bin\gcloud.cmd"
    r = subprocess.run([gcloud, "auth", "print-access-token"], capture_output=True, text=True, shell=True)
    return r.stdout.strip()

def fs_get(token, url):
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def fs_patch(token, doc_name, fields: dict):
    mask = "&".join(f"updateMask.fieldPaths={k}" for k in fields)
    url  = f"https://firestore.googleapis.com/v1/{doc_name}?{mask}"
    body = json.dumps({"fields": {
        k: {"stringValue": v} for k, v in fields.items()
    }}).encode()
    req = urllib.request.Request(url, data=body, method="PATCH", headers={
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })
    with urllib.request.urlopen(req) as r:
        return r.status

# ── Build lookup from spreadsheets ───────────────────────────────────────────
folder  = os.path.join(os.path.dirname(__file__), "..", "email_assets", "Leads")
lookup  = {}  # email -> {company_name, first_name_fix}

BAD_NAMES = {"nan", "none", "there", "", "null"}

def clean(v):
    s = str(v or "").strip()
    return s if s.lower() not in BAD_NAMES else None

# Bookstores: col0=BOOKSTORE NAME, col1=EMAIL
wb = openpyxl.load_workbook(os.path.join(folder, "Bookstores_BS.xlsx"))
for row in wb.active.iter_rows(min_row=2, values_only=True):
    co, em = clean(row[0]), clean(row[1])
    if em:
        lookup[em.lower()] = {"company_name": co, "first_name_fix": None}

# BookReviewBlogs: col0=EMAIL, col1=FIRST NAME
wb = openpyxl.load_workbook(os.path.join(folder, "BookReviewBlogs_BRB.xlsx"))
for row in wb.active.iter_rows(min_row=2, values_only=True):
    em, fn = clean(row[0]), clean(row[1])
    if em:
        lookup[em.lower()] = {"company_name": None, "first_name_fix": fn}

# BookReviewPodcasts: col0=FIRST NAME, col1=EMAIL, col2=PODCAST NAME
wb = openpyxl.load_workbook(os.path.join(folder, "BookReviewPodcasts_BRP.xlsx"))
for row in wb.active.iter_rows(min_row=2, values_only=True):
    fn, em, co = clean(row[0]), clean(row[1]), clean(row[2])
    if em:
        lookup[em.lower()] = {"company_name": co, "first_name_fix": fn}

# ChristianBlogs: col0=Email, col1=First Name, col2=Blog Name
wb = openpyxl.load_workbook(os.path.join(folder, "ChristianBlogs_ChrB.xlsx"))
for row in wb.active.iter_rows(min_row=2, values_only=True):
    em, fn, co = clean(row[0]), clean(row[1]), clean(row[2])
    if em:
        lookup[em.lower()] = {"company_name": co, "first_name_fix": fn}

# ChristianPodcasts: col0=Email, col1=Name, col2=Podcast Name
wb = openpyxl.load_workbook(os.path.join(folder, "ChristianPodcasts_ChrP.xlsx"))
for row in wb.active.iter_rows(min_row=2, values_only=True):
    em, fn, co = clean(row[0]), clean(row[1]), clean(row[2])
    if em:
        lookup[em.lower()] = {"company_name": co, "first_name_fix": fn}

print(f"Lookup: {len(lookup)} entries, "
      f"{sum(1 for v in lookup.values() if v['company_name'])} with company, "
      f"{sum(1 for v in lookup.values() if v['first_name_fix'])} with name fix")

# ── Fetch all outreach_leads from Firestore ───────────────────────────────────
token    = get_token()
all_docs = []
page_token = ""
while True:
    url = f"{BASE}/{COLL}?pageSize=300"
    if page_token:
        url += f"&pageToken={page_token}"
    data = fs_get(token, url)
    all_docs.extend(data.get("documents", []))
    page_token = data.get("nextPageToken", "")
    if not page_token:
        break
    time.sleep(0.1)

print(f"Fetched {len(all_docs)} outreach_leads from Firestore")

# ── Patch each doc ────────────────────────────────────────────────────────────
patched = skipped = errors = 0
token = get_token()  # refresh

for i, doc in enumerate(all_docs):
    if i % 500 == 0 and i > 0:
        token = get_token()   # refresh token every 500 ops
        print(f"  Progress: {i}/{len(all_docs)} — patched={patched} skipped={skipped} errors={errors}")

    f     = doc["fields"]
    email = f.get("email", {}).get("stringValue", "").lower()
    curr_fn = f.get("first_name", {}).get("stringValue", "")

    info  = lookup.get(email, {})
    co    = info.get("company_name")
    fn    = info.get("first_name_fix")

    updates = {}

    # Fix bad first_name
    if curr_fn.lower() in BAD_NAMES or curr_fn.lower() == "nan":
        if fn:
            updates["first_name"] = fn
        # else leave as "there" — email will say "Hey there"

    # Backfill good first_name even if current is ok
    if fn and curr_fn.lower() in BAD_NAMES:
        updates["first_name"] = fn

    # Add company_name if we have it
    if co and not f.get("company_name", {}).get("stringValue"):
        updates["company_name"] = co

    if not updates:
        skipped += 1
        continue

    try:
        fs_patch(token, doc["name"], updates)
        patched += 1
    except Exception as e:
        errors += 1
        if errors <= 5:
            print(f"  Error on {email}: {e}")

print(f"\nDone: patched={patched}  skipped={skipped}  errors={errors}")
